import openpyxl
import threading

from datetime import datetime
from string import ascii_uppercase
from simple_history.utils import bulk_create_with_history, bulk_update_with_history

from django.contrib import messages
from django.core.mail import EmailMultiAlternatives
from django.db.models import Q, Max
from django.db.models.functions import Length
from django.shortcuts import render, redirect
from django.template.loader import get_template
from django.utils.translation import gettext as _
from django.views.generic import TemplateView
from django.views.generic.edit import FormMixin
from django.urls import reverse_lazy

from .models import (Bodega, Estante, Nivel, Posicion, Caja, Cliente, Moneda,
    Producto, Oficina, Credito, Tomo)
from .forms import (Busqueda, GeneraEstructura, GeneraEtiquetas_Form, 
    CargaCreditos_Form, IngresoTomo_Form, EgresoTomo_Form, Bodega_From, 
    TrasladoTomos_Form, SalidaTomos_Form)
from usuarios.views_base import (ListView_Login, DetailView_Login, TemplateView_Login, 
    CreateView_Login, UpdateView_Login, DeleteView_Login, FormView_Login)

# Create your views here.
class Inicio_Template(TemplateView):
    template_name = 'documentos/index.html'
    extra_context = {
        'title': 'Documentos',
    }


class Bodega_ListView(ListView_Login):
    permission_required = 'documentos.view_bodega'
    model = Bodega
    paginate_by = 15
    ordering = ['-vigente', 'nombre']
    extra_context = {
        'title': _('Bodegas'),
        'opciones': {
            'etiqueta': _('Opciones'),
            'ver': _('Ver'),
            'editar': _('Editar'),
            'nuevo': _('Nuevo'),
        },
        'botones': {
            'buscar': _('Buscar'),
            'limpiar': _('Limpiar'),
        },
        'mensaje_vacio': _('No hay "Bodegas" registradas'),
    }

    def get_context_data(self, *args, **kwargs):
        busqueda = self.request.GET.get('valor')

        context = super().get_context_data(*args, **kwargs)
        context['url_lista'] = Bodega.list_url()
        if busqueda:
            context['object_list'] = Bodega.objects.filter(nombre__icontains=busqueda).order_by('nombre')
            context['form'] = Busqueda(self.request.GET)
        else:
            context['form'] = Busqueda()
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_superuser:
            return queryset
        else:    
            return queryset.filter(
                Q(personal=self.request.user)
                |Q(encargado=self.request.user)).distinct()

class Bodega_DetailView(FormMixin, DetailView_Login):
    permission_required = 'documentos.view_bodega'
    model = Bodega
    form_class = GeneraEstructura
    extra_context = {
        'title': _('Bodega'),
        'sub_titulo': {
            'genera': _('Genera Estructura'),
            'estructura': _('Estructura'),
        },
        'botones': {
            'generar': _('Generar'),
        },
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['opciones']={
            'etiqueta': _('Opciones'),
            'editar': _('Editar'),
            'cajas_inhabilitadas': _('Cajas Inhabilitadas'),
        }
        queryset = Nivel.objects.filter(estante__bodega=self.object)\
            .order_by(Length('estante__codigo'), 'estante__codigo', 'numero')\
            .prefetch_related('estante', 'estante__bodega')
        context['estructura']={
            'tooltip_head': _('Estante'),
            'tooltip_body': _('Nivel'),
            'estructura': queryset,
            'max_columnas': queryset.aggregate(Max('numero'))['numero__max'],
        }
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_superuser:
            return queryset
        else:    
            return queryset.filter(
                Q(personal=self.request.user)
                |Q(encargado=self.request.user)).distinct()

    def get_success_url(self, *args, **kwargs):
        return self.object.view_url()

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        if 'Generar' in request.POST:
            form = self.get_form()
            if form.is_valid():
                return self.form_valid(form)
            else: 
                return self.form_invalid(form)
    
    def form_valid(self, form, *args, **kwargs):
        e = form.cleaned_data['estantes']
        n = form.cleaned_data['niveles']
        p = form.cleaned_data['posiciones']
        c = form.cleaned_data['cajas']
        self._genera_estructura(e, n, p, c)

        return super().form_valid(form)

    def _genera_estructura(self, pEstantes, pNiveles, pPosiciones, pCajas):
        self._genera_estructura_estantes(pEstantes)
        self._genera_estructura_niveles(pNiveles)
        self._genera_estructura_posiciones(pPosiciones)
        self._genera_estructura_cajas(pCajas)
    
    def _genera_estructura_cajas(self, pCaja):
        inserts = []
        cajas = Caja.objects.filter(posicion__nivel__estante__bodega=self.object)
        for posicion in Posicion.objects.filter(nivel__estante__bodega=self.object, vigente=True):
            for i in range(1, pCaja+1):
                if not cajas.filter(numero=i, posicion=posicion):
                    inserts.append(Caja(numero=i, posicion=posicion))

        bulk_create_with_history(inserts, Caja, batch_size=1500)

    def _genera_estructura_posiciones(self, pPosicion):
        inserts = []
        posiciones = Posicion.objects.filter(nivel__estante__bodega=self.object)
        for nivel in Nivel.objects.filter(estante__bodega=self.object, vigente=True):
            for i in range(1, pPosicion+1):
                if not posiciones.filter(numero=i, nivel=nivel):
                    inserts.append(Posicion(numero=i, nivel=nivel))

        Posicion.objects.bulk_create(inserts)

    def _genera_estructura_niveles(self, pNiveles):
        inserts = []
        niveles = Nivel.objects.filter(estante__bodega=self.object)
        for estante in Estante.objects.filter(bodega=self.object, vigente=True):
            for i in range(1, pNiveles+1):
                if not niveles.filter(numero=i, estante=estante):
                    inserts.append(Nivel(numero=i, estante=estante))

        Nivel.objects.bulk_create(inserts)

    def _genera_estructura_estantes(self, pEstante):
        inserts = []
        estantes = Estante.objects.filter(bodega=self.object.id)

        for codigo in self._codigo_estante(pEstante):
            if not estantes.filter(codigo=codigo):
                inserts.append(Estante(codigo=codigo, bodega=self.object))

        Estante.objects.bulk_create(inserts)

    def _codigo_estante(self, numero=0):
        codigos = []
        for i in range(numero):
            c1 = i // 26
            c2 = i - c1*26
            codigos.append('{}{}'.format('' if i<26 else ascii_uppercase[c1-1],ascii_uppercase[c2]))
            i+=1
        return codigos

class Bodega_CreateView(CreateView_Login):
    permission_required = 'documentos.add_bodega'
    model = Bodega
    form_class = Bodega_From
    template_name = 'documentos/form.html'
    extra_context = {
        'title': _('Nueva Bodega'),
        'botones': {
            'guardar': _('Guardar'),
            'cancelar': _('Cancelar'),
        },
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['list_url'] = Bodega.list_url()
        return context

class Bodega_UpdateView(UpdateView_Login):
    permission_required = 'documentos.change_bodega'
    model = Bodega
    form_class = Bodega_From
    template_name = 'documentos/form.html'
    extra_context = {
        'title': _('Modificar Bodega'),
        'botones': {
            'guardar': _('Guardar'),
            'cancelar': _('Cancelar'),
        },
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['list_url'] = Bodega.list_url()
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_superuser:
            return queryset
        else:
            return queryset.filter(
                Q(personal=self.request.user)
                |Q(encargado=self.request.user)).distinct()

class Bodega_DeleteView(DeleteView_Login):
    permission_required = 'documentos.delete_bodega'
    model = Bodega
    template_name = 'documentos/confirmation_form.html'
    extra_context = {
        'title': _('Cambiar Estado Bodega'),
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['list_url'] = Bodega.list_url()
        context['botones']={
            'cancelar': _('Cancelar'),
        }
        context['mensajes']={
            'confirmacion': _(f'¿Quiere {self.object.get_accion()} el elemento indicado?'),
        }
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_superuser:
            return queryset
        else:
            return queryset.filter(
                Q(personal=self.request.user)
                |Q(encargado=self.request.user)).distinct()

    def get_success_url(self, *args, **kwargs):
        return super().get_success_url(self.object.vigente)


class Estante_DetailView(DetailView_Login):
    permission_required = 'documentos.view_estante'
    model = Estante
    extra_context = {
        'title': _('Estante'),
        'sub_titulo': {
            'estructura': _('Estructura'),
        },
        'opciones':{
            'etiqueta':_('Opciones'),
            'etiquetas':_('Etiquetas'),
        },
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        
        queryset = Posicion.objects.filter(nivel__estante=self.object)\
            .order_by('nivel__numero', 'numero')\
            .prefetch_related('nivel', 'nivel__estante', 'nivel__estante__bodega')

        context['estructura']={
            'tooltip_head': _('Nivel'),
            'tooltip_body': _('Posición'),
            'estructura': queryset,
            'max_columnas': queryset.aggregate(Max('numero'))['numero__max'],
        }
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_superuser:
            return queryset
        else:    
            return queryset.filter(
                Q(bodega__personal=self.request.user)
                |Q(bodega__encargado=self.request.user)).distinct()

class Estante_Etiqueta(DetailView_Login):
    permission_required = 'documentos.label_estante'
    template_name = 'documentos/etiqueta.html'
    model = Caja

    def get(self, request, *args, **kwargs):
        context = {
            'title': _('Etiqueta'),
            'object': Estante.objects.get(pk=kwargs['pk'])
        }
        arreglo = {}
        cajas = list(Caja.objects.filter(posicion__nivel__estante__id=self.kwargs['pk']))
        for caja in cajas:
            arreglo[caja] = 'mostrar'
        return render(request, self.template_name, {'arreglo': arreglo, 'context': context})


class Nivel_DetailView(DetailView_Login):
    permission_required = 'documentos.view_nivel'
    model = Nivel
    extra_context = {
        'title': _('Nivel'),
        'sub_titulo': {
            'estructura': _('Estructura'),
        },
        'opciones':{
            'etiqueta':_('Opciones'),
            'etiquetas':_('Etiquetas'),
        },
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        
        queryset = Caja.objects.filter(posicion__nivel=self.object)\
            .order_by('posicion__numero', 'numero')\
            .prefetch_related('posicion__nivel', 'posicion__nivel__estante', 
                'posicion__nivel__estante__bodega')

        context['estructura']={
            'tooltip_head': _('Posición'),
            'tooltip_body': _('Caja'),
            'estructura': queryset,
            'max_columnas': queryset.aggregate(Max('numero'))['numero__max'],
        }
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_superuser:
            return queryset
        else:    
            return queryset.filter(
                Q(estante__bodega__personal=self.request.user)
                |Q(estante__bodega__encargado=self.request.user)).distinct()

class Nivel_Etiqueta(DetailView_Login):
    permission_required = 'documentos.label_nivel'
    template_name = 'documentos/etiqueta.html'
    model = Caja
    
    def get(self, request, *args, **kwargs):
        context = {
            'title': _('Etiqueta'),
            'object': Nivel.objects.get(pk=kwargs['pk'])
        }
        arreglo = {}
        cajas = list(Caja.objects.filter(posicion__nivel__id=self.kwargs['pk']))
        for caja in cajas:
            arreglo[caja] = 'mostrar'
        return render(request, self.template_name, {'arreglo': arreglo, 'context': context})


class Posicion_DetailView(DetailView_Login):
    permission_required = 'documentos.view_posicion'
    model = Posicion
    extra_context = {
        'title': _('Posición'),
        'sub_titulo': {
            'estructura': _('Estructura'),
        },
        'opciones':{
            'etiqueta':_('Opciones'),
            'etiquetas':_('Etiquetas'),
        },
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        
        queryset = Caja.objects.filter(posicion=self.object)\
            .order_by('numero')\
            .prefetch_related('posicion', 'posicion__nivel', 
                'posicion__nivel__estante', 
                'posicion__nivel__estante__bodega')

        context['estructura']={
            'tooltip_head': _('Caja'),
            'tooltip_body': _('Tomo'),
            'estructura': queryset,
        }
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_superuser:
            return queryset
        else:
            return queryset.filter(
                Q(nivel__estante__bodega__personal=self.request.user)
                | Q(nivel__estante__bodega__encargado=self.request.user)).distinct()

class Posicion_Etiqueta(DetailView_Login):
    permission_required = 'documentos.label_posicion'
    template_name = 'documentos/etiqueta.html'
    model = Caja
    
    def get(self, request, *args, **kwargs):
        context = {
            'title': _('Etiqueta'),
            'object': Posicion.objects.get(pk=kwargs['pk'])
        }
        arreglo = {}
        cajas = list(Caja.objects.filter(posicion__id=self.kwargs['pk']))
        for caja in cajas:
            arreglo[caja] = 'mostrar'
        return render(request, self.template_name, {'arreglo': arreglo, 'context': context})


class Caja_ListView(ListView_Login):
    permission_required = 'documentos.view_caja'
    model = Caja
    paginate_by = 15
    ordering = ['posicion__nivel__estante__codigo', 'posicion__nivel__numero', 'posicion__numero', 'numero']
    extra_context = {
        'title': _('Cajas Inhabilitadas'),
        'etiquetas': {
            'caja': _('Caja'),
        },
        'opciones': {
            'etiqueta': _('Opciones'),
            'habilitar': _('Habilitar'),
        },
        'mensaje_vacio': _('No hay "Cajas" inhabilitadas'),
    }

    def get_queryset(self):
        queryset = super().get_queryset().filter(vigente=False)
        if self.request.user.is_superuser:
            return queryset
        else:
            return queryset.filter(
                Q(posicion__nivel__estante__bodega__personal=self.request.user)
                |Q(posicion__nivel__estante__bodega__encargado=self.request.user)).distinct()

class Caja_DetailView(DetailView_Login):
    permission_required = 'documentos.view_caja'
    model = Caja
    extra_context = {
        'title': _('Caja'),
        'sub_titulo': {
            'tomos': _('Tomos')
        },
        'opciones': {
            'etiqueta':_('Opciones'),
            'etiquetas':_('Etiquetas'),
        },
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        queryset = Tomo.objects.filter(caja=self.object)\
            .order_by('-fecha_modificacion')\
            .prefetch_related('credito')
        context['estructura']={
            'tooltip_head': _('Tomo'),
            'estructura': queryset,
        }
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_superuser:
            return queryset
        else:
            return queryset.filter(
                Q(posicion__nivel__estante__bodega__personal=self.request.user)
                |Q(posicion__nivel__estante__bodega__encargado=self.request.user)).distinct()

class Caja_DeleteView(DeleteView_Login):
    permission_required = 'documentos.delete_caja'
    model = Caja
    template_name = 'documentos/confirmation_form.html'
    extra_context = {
        'title': _('Cambiar Estado de Caja'),
        'botones': {
            'cancelar': _('Cancelar'),
        },
        'list_url': reverse_lazy('documentos:index'),
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['mensajes']={
            'confirmacion': _(f'¿Quiere {self.object.get_accion()} el elemento indicado?'),
        }
        return context

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_superuser:
            return queryset
        else:
            return queryset.filter(
                Q(posicion__nivel__estante__bodega__personal=self.request.user)
                |Q(posicion__nivel__estante__bodega__encargado=self.request.user)).distinct()

    def get_success_url(self, *args, **kwargs):
        return self.object.view_url()

class Caja_Etiqueta(DetailView_Login):
    permission_required = 'documentos.label_caja'
    template_name = 'documentos/etiqueta.html'
    model = Caja
    
    def get(self, request, *args, **kwargs):
        caja = Caja.objects.get(id=self.kwargs['pk'])
        context = {
            'title': _('Etiqueta'),
            'object': caja
        }
        arreglo = {}
        arreglo[caja] = 'mostrar'
        return render(request, self.template_name, {'arreglo': arreglo, 'context': context})


class CargaMasiva_Form(FormView_Login):
    permission_required = 'documentos.load_credito'
    form_class = CargaCreditos_Form
    template_name = 'documentos/form_loadfile.html'
    success_message = _('Se cargaron los registros exitosamente')
    extra_context = {
        'title': _('Carga de Créditos'),
        'botones': {
            'guardar': _('Cargar'),
            'cancelar': _('Cancelar'),
        }
    }

    def get_success_url(self):
        return reverse_lazy('documentos:index')

    def post(self, request, *args, **kwargs):
        if 'Cargar' in request.POST:
            form = self.get_form()
            if form.is_valid():
                return self.form_valid(form)
            else: 
                return self.form_invalid(form)

    def form_valid(self, form):
        archivo = form.cleaned_data['archivo']
        libro = openpyxl.load_workbook(archivo)
        _insert_datos(datos_excel(libro[libro.sheetnames[0]]))
        return super().form_valid(form)

class Credito_DetailView(DetailView_Login):
    permission_required = 'documentos.view_credito'
    model = Credito
    extra_context = {
        'title': _('Credito'),
        'sub_titulo': {
            'tomos': _('Tomos'),
        },
        'etiquetas':{
            'tomo':_('Cant. de Tomos'),
        },
        'opciones':{
            'etiqueta':_('Opciones'),
            'etiquetas':_('Etiquetas'),
            'escaneado': _('Escaneado'),
            'agregar_tomo':_('Agregar tomo'),
            'remover_tomo':_('Remover tomo'),
            'extraer':_('Extraer'),
            'guardar':_('Guardar'),
        },
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['egreso_form']=EgresoTomo_Form()
        context['tomos'] = Tomo.objects.filter(credito=self.object, vigente=True)\
            .order_by('numero').prefetch_related('caja', 'caja__posicion', 
                'caja__posicion__nivel', 'caja__posicion__nivel__estante', 
                'caja__posicion__nivel__estante__bodega',
                'caja__posicion__nivel__estante__bodega__personal')
        return context

class Credito_Etiqueta(DetailView_Login):
    permission_required = 'documentos.label_credito'
    template_name = 'documentos/etiqueta_tomo.html'
    model = Tomo
    
    def get(self, request, *args, **kwargs):
        context = {
            'title': _('Etiqueta'),
            'object': Credito.objects.get(pk=kwargs['pk'])
        }
        arreglo = {}
        tomos = list(Tomo.objects.filter(credito__id=self.kwargs['pk'], vigente=True).order_by('numero'))
        for tomo in tomos:
            arreglo[tomo] = 'mostrar'
        return render(request, self.template_name, {'arreglo': arreglo, 'context': context})


class Tomo_Ingreso(FormView_Login):
    permission_required = 'documentos.change_tomo'
    form_class = IngresoTomo_Form
    template_name = 'documentos/form.html'
    extra_context = {
        'title': _('Ingreso a Bodega'),
        'botones': {
            'guardar': _('Guardar'),
            'cancelar': _('Cancelar'),
        },
        'list_url': reverse_lazy('documentos:index'),
    }
    success_url = reverse_lazy('documentos:index')
    success_message ='' # se remueve el mensaje default agregado en formview_login

    def post(self, request, *args, **kwargs):
        if 'Guardar' in request.POST:
            form = self.get_form()
            if form.is_valid():
                return self.form_valid(form)
            else: 
                return self.form_invalid(form)
    
    def form_valid(self, form, *args, **kwargs):
        tomo = form.cleaned_data['tomo']
        caja = form.cleaned_data['caja']
        comentario = form.cleaned_data['comentario']

        try:
            tomo_qs = Tomo.objects.get(credito__numero=tomo[0], numero=tomo[1])
            caja_qs = Caja.objects.get(posicion__nivel__estante__bodega__codigo=caja[0].upper(), 
                posicion__nivel__estante__codigo=caja[1].upper(), posicion__nivel__numero=caja[2], 
                posicion__numero=caja[3], numero=caja[4])
            
            if not self.request.user in caja_qs.posicion.nivel.estante.bodega.personal.all():
                messages.warning(self.request, _('Usuario no puede asignar expedientes en esa caja'))
            elif not caja_qs.vigente:
                messages.warning(self.request, _('La caja no se encuentra habilitada'))
            elif not caja_qs.posicion.nivel.estante.bodega.vigente:
                messages.warning(self.request, _('La bodega no se encuentra habilitada'))
            elif not tomo_qs.caja:
                tomo_qs.comentario, tomo_qs.vigente = comentario, True
                tomo_qs.caja, tomo_qs.usuario = caja_qs, self.request.user
                tomo_qs._change_reason, tomo_qs._history_user = 'Tomo_Ingreso', self.request.user
                tomo_qs.save()
                messages.success(self.request, _('Se guardo el tomo ')+f'{tomo[0]}-{tomo[1]}')

                return redirect(tomo_qs.credito.view_url())
            else:
                messages.warning(self.request, _('El tomo se encuentra asignado a: ')+f'{tomo_qs.caja}')
        except:
            messages.warning(self.request, _('Tomo o caja no encontrado ')+f'{tomo[0]}-{tomo[1]}')
        return super().form_valid(form)

class Tomo_Etiqueta(DetailView_Login):
    permission_required = 'documentos.label_tomo'
    template_name = 'documentos/etiqueta_tomo.html'
    model = Tomo

    def get(self, request, *args, **kwargs):
        context = {
            'title': _('Etiqueta'),
            'object': Credito.objects.get(tomo_credito__pk=self.kwargs['pk'])
        }
        arreglo = {}    
        tomos = list(Tomo.objects.filter(id=self.kwargs['pk']))
        for tomo in tomos:
            arreglo[tomo] = 'mostrar'
        return render(request, self.template_name, {'arreglo': arreglo, 'context': context})

class Tomo_Template(TemplateView_Login):
    permission_required = 'documentos.change_tomo'
    template_name = 'documentos/tomo_list.html'
    extra_context = {
        'title': 'Salida de Tomos',
        'opciones': {
            'etiqueta': _('Opciones'),
            'eliminar': _('Eliminar'),
            'trasladar': _('Trasladar'),
            'egresar': _('Egresar'),
        },
        'sub_titulo': {
            'traslado':_('Traslado de bodega'),
            'egreso': _('Egreso de bodega'),
        },
        'mensaje_vacio':_('No hay tomos para enviar')
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        tomos = self.request.session.get('extraer_tomos', [])
        tomo_qs = Tomo.objects.filter(id__in=tomos)\
            .order_by('credito__numero', 'numero')
        context['object_list'] = tomo_qs
        context['traslado_form'] = TrasladoTomos_Form()
        context['egreso_form'] = SalidaTomos_Form()
        return context


def buscar_credito(request):
    numero = request.GET['numero'].replace(' ','').split('-')
    credito = Credito.objects.filter(numero=numero[0])
    if credito:
        return redirect(credito[0].view_url())
    elif len(numero)>0:
        messages.warning(request, _('No se encontró el crédito: ')+numero[0])

    return redirect(reverse_lazy('documentos:index'))
    
def operaciones_tomo(request, pk=None):
    if 'escaneado.x' in request.POST:
        ''' Actualiza el Crédito a Escaneado '''
        credito = Credito.objects.get(id=request.POST['credito'])
        credito.escaneado = True
        credito._history_user = request.user
        credito._change_reason = 'operaciones_tomo > escaneado '
        credito.save()
    elif 'agregar.x' in request.POST:
        ''' Crea/Habilita un tomo '''
        credito = Credito.objects.get(id=request.POST['credito'])
        tomos = Tomo.objects.filter(credito=credito)
    
        if not tomos:
            tomo = Tomo(numero=1, credito=credito, comentario='Tomo habilitado', 
                usuario=request.user)
            tomo._change_reason = 'operaciones_tomo > agregar (unico nuevo)'
            tomo._history_user = request.user
            tomo.save()
        elif tomos.filter(vigente=False):
            #tomo minimo inhabilitado
            tomo = tomos.filter(vigente=False).order_by('numero')[0]
            tomo.usuario, tomo.vigente = request.user, True
            tomo.comentario = 'Tomo habilitado'
            tomo._change_reason, tomo._history_user = 'operaciones_tomo > agregar (habilita existente)', request.user
            tomo.save()
        else:
            num = tomos.aggregate(Max('numero'))['numero__max']+1
            tomo = Tomo(numero=num, credito=credito, comentario='Tomo habilitado', 
                usuario=request.user)
            tomo._change_reason = 'operaciones_tomo > agregar (agrega nuevo)'
            tomo._history_user = request.user
            tomo.save()
    elif 'remover.x' in request.POST:
        ''' Ihabilita un tomo '''
        credito = Credito.objects.get(id=request.POST['credito'])
        tomos = Tomo.objects.filter(credito=credito)
    
        if not tomos or not tomos.filter(vigente=True):
            messages.warning(request, _('No hay tomos para deshabilitar en el crédito ')+ credito.numero)
        else:
            #tomo máximo habilitado
            tomo = tomos.filter(vigente=True).order_by('-numero')[0]
            if tomo.caja:
                messages.warning(request, _('El tomo se encuentra ingresado en una caja '))
            else:
                tomo.usuario, tomo.vigente = request.user, False
                tomo.comentario='Tomo inhabilitado'
                tomo._change_reason, tomo._history_user = 'operaciones_tomo > remover', request.user
                tomo.save()
    elif 'agregar' in request.POST:
        ''' Agrega tomos a a lista que se extraera de la bodega '''
        extraer_tomos=request.session['extraer_tomos'] if 'extraer_tomos' in request.session else []
        tomoid = request.POST['tomo-id']
        tomo = request.POST['tomo'].replace(' ', '').split('-')

        try: 
            if len(tomo)==2:
                t = Tomo.objects.get(credito__numero=tomo[0], numero=int(tomo[1]))
                if t and str(t.id)==tomoid:
                    if not tomoid in extraer_tomos:
                        extraer_tomos.append(tomoid)
                        request.session['extraer_tomos']=extraer_tomos
                        messages.success(request, _('Tomo agregado a la lista'))
                    else:
                        messages.warning(request, _('Tomo agregado previamente'))
                else:
                    messages.warning(request, _('Tomo no corresponde al seleccionado'))
            else:
                messages.warning(request, _('Tomo mal ingresado (longitud)'))
        except:
            messages.warning(request, _('Tomo mal ingresado o no existe'))
        finally:
            return redirect(Tomo.objects.get(id=tomoid).credito.view_url())   
    else: #'quitar' in request.POST:
        ''' Elimina tomos del listado a trasladar/egresar '''
        try: 
            extraer_tomos = request.session['extraer_tomos']
            extraer_tomos.remove(str(pk))
            request.session['extraer_tomos'] = extraer_tomos
        finally:
            return redirect(Tomo.envio_url())
    return redirect(credito.view_url())
    
def salida_tomo(request):
    if 'trasladar' in request.POST:
        try:
            form = TrasladoTomos_Form(request.POST)
            if form.is_valid():
                bodega = form.cleaned_data['bodega_envio']
                comentario = f'Traslado a {bodega}\n{form.cleaned_data["comentario"]}'
                tomos = Tomo.objects.filter(id__in=request.session['extraer_tomos'])
                tomos.update(comentario=comentario, caja=None, usuario=request.user)
                for tomo in tomos : tomo._change_reason, tomo._history_user = 'salida_tomo > trasladar', request.user
                bulk_update_with_history(tomos, Tomo, ['comentario', 'caja'], batch_size=1500)
                del request.session['extraer_tomos']
                messages.success(request, _('Tomos egresados por traslado'))
                
                context = {
                    'title': _('Traslado'),
                    'message': _(f'Se han enviado, a {bodega}, los siguientes tomos:'),
                    'object_list': {tomo for tomo in tomos},
                }
                correo = crea_correo('Traslado de Expedientes', request.user.email, [bodega.encargado.email], 'mails/egresos.html', context)
                envia_correo(correo)
        finally:
            return redirect(Tomo.envio_url())
    elif 'egresar' in request.POST:
        try:
            form = SalidaTomos_Form(request.POST)
            if form.is_valid():
                correo_form = form.cleaned_data["correo"]
                print(f"correo>>>{correo_form}")
                comentario_final = f'Fecha: \t\t{datetime.today().strftime("%d-%m-%Y")}\n'
                comentario_final += f'Codigo: \t{form.cleaned_data["codigo"]}\n'
                comentario_final += f'Nombre: \t{form.cleaned_data["nombre"]}\n'
                comentario_final += f'Extension: \t{form.cleaned_data["extension"]}\n'
                comentario_final += f'Correo: \t{correo_form}\n'
                comentario_final += f'Gerencia: \t{form.cleaned_data["gerencia"]}\n'
                comentario_final += f'Comentario: \t{form.cleaned_data["comentario"]}'

                tomos = Tomo.objects.filter(id__in=request.session['extraer_tomos'])
                tomos.update(comentario=comentario_final, caja=None, usuario=request.user)
                for tomo in tomos : tomo._change_reason, tomo._history_user = 'salida_tomo > egresar', request.user
                bulk_update_with_history(tomos, Tomo, ['comentario', 'caja'], batch_size=1500)
                del request.session['extraer_tomos']
                messages.success(request, _('Tomos egresados por solicitud'))
                
                context = {
                    'title': _('Egreso por Solicitud'),
                    'message': _(f'Se han entregado, los siguientes tomos:'),
                    'object_list': {tomo for tomo in tomos},
                    'comentario': comentario_final,
                }
                correo = crea_correo('Egreso por solicitud', request.user.email, [request.user.email,correo_form], 'mails/egresos.html', context)
                envia_correo(correo)
        finally:
            return redirect(Tomo.envio_url())



##########################################################################
#
##########################################################################
def crea_correo(subject, from_email, to_email, template, context):
    template = get_template(template)
    content = template.render(context)

    mail = EmailMultiAlternatives(
        subject=subject,
        body='',
        from_email=from_email,
        to=to_email,
        cc=[],
        )

    mail.attach_alternative(content, 'text/html')
    return mail

def envia_correo(mail):
    ''' Envío asincrono de correos '''
    thread = threading.Thread(
        mail.send(fail_silently=False)
    )
    thread.start()
    
##########################################################################
#
##########################################################################
def datos_excel(hoja):    
    orden = {}
    for i in range(len(hoja[1])):
        orden[hoja[1][i].value]=i

    clientes = []
    oficinas = []
    monedas = []
    productos = []
    creditos = []
    for fila in hoja.iter_rows(min_row=2):
        mis     = fila[orden['Cod_Cliente']].value
        cliente = fila[orden['Cliente']].value
        if not any(mis in sublist for sublist in clientes):
            clientes.append([mis, cliente])

        cod_ofi = fila[orden['Cod_ofi']].value
        oficina = fila[orden['Oficina']].value 
        if not any(cod_ofi in sublist for sublist in oficinas):
            oficinas.append([cod_ofi, oficina])

        moneda  = fila[orden['Moneda']].value
        if not moneda in monedas:
            monedas.append(moneda)

        producto = fila[orden['Producto']].value
        if not producto in productos:
            productos.append(producto)

        credito = fila[orden['Credito']].value
        fecha   = datetime.strptime(fila[orden['Fecha_Ini']].value, "%d/%m/%Y").strftime("%Y-%m-%d")
        monto   = fila[orden['Monto']].value
        if not any(credito in sublist for sublist in creditos):
            creditos.append([credito, fecha, monto, mis, cod_ofi, moneda, producto])
        
    return {'clientes': clientes, 'oficinas': oficinas, 'monedas': monedas, 
        'productos': productos,'creditos': creditos}

def _insert_datos(datos):
    clientes = _insert_clientes(datos['clientes'])
    oficinas = _insert_oficinas(datos['oficinas'])
    monedas = _insert_monedas(datos['monedas'])
    productos = _insert_productos(datos['productos'])
    creditos = _insert_creditos(datos['creditos'])

def _insert_clientes(datos):
    clientes = []
    queryset = Cliente.objects.filter(codigo__in=[int(dato[0]) for dato in datos])
    for dato in datos:
        if not queryset.filter(codigo=int(dato[0])):
            clientes.append(Cliente(codigo=int(dato[0]), nombre=dato[1]))
    Cliente.objects.bulk_create(clientes)

def _insert_oficinas(datos):
    oficinas = []
    queryset = Oficina.objects.filter(numero__in=[int(dato[0]) for dato in datos])
    for dato in datos:
        if not queryset.filter(numero=int(dato[0])):
            oficinas.append(Oficina(numero=int(dato[0]), descripcion=dato[1]))
    Oficina.objects.bulk_create(oficinas)

def _insert_monedas(datos):
    monedas = []
    queryset = Moneda.objects.filter(descripcion__in=datos)
    for dato in datos:
        if not queryset.filter(descripcion=dato):
            monedas.append(Moneda(descripcion=dato))
    Moneda.objects.bulk_create(monedas)

def _insert_productos(datos):
    productos = []
    queryset = Producto.objects.filter(descripcion__in=datos)
    for dato in datos:
        if not queryset.filter(descripcion=dato):
            productos.append(Producto(descripcion=dato))
    Producto.objects.bulk_create(productos)

def _insert_creditos(datos):
    queryset = Credito.objects.filter(numero__in=[dato[0] for dato in datos])
    clientes = Cliente.objects.filter(codigo__in=[int(dato[3]) for dato in datos])
    oficinas = Oficina.objects.filter(numero__in=[int(dato[4]) for dato in datos])
    monedas = Moneda.objects.filter(descripcion__in=[dato[5] for dato in datos])
    productos = Producto.objects.filter(descripcion__in=[dato[6] for dato in datos])

    creditos = []
    for dato in datos:
        if not queryset.filter(numero=dato[0]):
            creditos.append(Credito(numero=dato[0], fecha_concesion=dato[1], 
                monto=dato[2], cliente=clientes.get(codigo=int(dato[3])), 
                moneda=monedas.get(descripcion=dato[5]), 
                oficina=oficinas.get(numero=int(dato[4])), 
                producto=productos.get(descripcion=dato[6])))
    bulk_create_with_history(creditos, Credito, batch_size=1500)