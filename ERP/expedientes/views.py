import openpyxl
import threading

from datetime import datetime
from itertools import product
from string import ascii_uppercase

from django.conf import settings
from django.contrib import messages
from django.core.mail import EmailMultiAlternatives
from django.db.models import Count, Max, Min
from django.db.models.functions import Length
from django.shortcuts import render, redirect
from django.template.loader import get_template
from django.utils.translation import gettext as _
from django.views.generic import TemplateView, RedirectView
from django.views.generic.edit import FormMixin
from django.urls import reverse_lazy

from .models import (Bodega, Estante, Nivel, Posicion, Caja, Cliente, Moneda,
    Producto, Oficina, Credito, Tomo)
from .forms import (Busqueda, GeneraEstructura, GeneraEtiquetas_Form, 
    CargaCreditos_Form, IngresoTomo_Form, EgresoTomo_Form, Bodega_From, 
    TrasladoTomos_Form, SalidaTomos_Form)
from usuarios.views_base import (ListView_Login, DetailView_Login, TemplateView_Login, 
    CreateView_Login, UpdateView_Login, DeleteView_Login, FormView_Login)

# Create your views here.
class Inicio_Template(TemplateView):
    template_name = 'expedientes/index.html'
    extra_context = {
        'title': 'Expedientes',
    }


class Bodega_ListView(ListView_Login):
    permission_required = 'expedientes.view_bodega'
    model = Bodega
    paginate_by = 15
    ordering = ['-vigente', 'nombre']
    extra_context = {
        'title': _('Bodegas'),
        'opciones': {
            'etiqueta': _('Opciones'),
            'ver': _('Ver'),
            'editar': _('Editar'),
            'inhabilitar': _('Inhabilitar'),
            'habilitar': _('Habilitar'),
            'nuevo': _('Nuevo'),
        },
        'botones': {
            'buscar': _('Buscar'),
            'limpiar': _('Limpiar'),
        },
        'mensaje_vacio': _('No hay "Bodegas" registradas'),
    }

    def get_context_data(self, *args, **kwargs):
        busqueda = self.request.GET.get('valor')

        context = super().get_context_data(*args, **kwargs)
        context['url_lista'] = Bodega.list_url()
        if busqueda:
            context['object_list'] = Bodega.objects.filter(nombre__icontains=busqueda).order_by('nombre')
            context['form'] = Busqueda(self.request.GET)
        else:
            context['form'] = Busqueda()
        return context

class Bodega_DetailView(FormMixin, DetailView_Login):
    permission_required = 'expedientes.view_bodega'
    model = Bodega
    form_class = GeneraEstructura
    extra_context = {
        'title': _('Bodega'),
        'sub_titulo': {
            'genera': _('Genera Estructura'),
            'estructura': _('Estructura'),
        },
        'botones': {
            'generar': _('Generar'),
        }
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['opciones']={
            'etiqueta': _('Opciones'),
            'editar': _('Editar'),
            'cajas_inhabilitadas': _('Cajas Inhabilitadas'),
            'accion': _('Inhabilitar') if self.object.vigente else _('Habilitar'),
            'accion_tag': 'danger' if self.object.vigente else 'success',
        }
        queryset = Estante.objects.filter(bodega=self.object).order_by(Length('codigo'), 'codigo')
        context['estructura']={
            'estantes': queryset,
            'no_columnas': queryset.count(),
            'tooltip_head': _('Estante'),
            'tooltip_body': _('Nivel'),
            'niveles': Nivel.objects.filter(estante__bodega=self.object).order_by('numero', Length('estante__codigo'), 'estante__codigo')
        }
        return context

    def get_success_url(self, *args, **kwargs):
        return self.object.view_url()

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        if 'Generar' in request.POST:
            form = self.get_form()
            if form.is_valid():
                return self.form_valid(form)
            else: 
                return self.form_invalid(form)
    
    def form_valid(self, form, *args, **kwargs):
        e = form.cleaned_data['estantes']
        n = form.cleaned_data['niveles']
        p = form.cleaned_data['posiciones']
        c = form.cleaned_data['cajas']
        self._genera_estructura(e, n, p, c)

        return super().form_valid(form)

    def _genera_estructura(self, pEstantes, pNiveles, pPosiciones, pCajas):
        self._genera_estructura_estantes(pEstantes)
        self._genera_estructura_niveles(pNiveles)
        self._genera_estructura_posiciones(pPosiciones)
        self._genera_estructura_cajas(pCajas)
    
    def _genera_estructura_cajas(self, pCaja):
        inserts = []
        cajas = Caja.objects.filter(posicion__nivel__estante__bodega=self.object)
        for posicion in Posicion.objects.filter(nivel__estante__bodega=self.object, vigente=True):
            for i in range(1, pCaja+1):
                if not cajas.filter(numero=i, posicion=posicion):
                    inserts.append(Caja(numero=i, posicion=posicion))

        Caja.objects.bulk_create(inserts)

    def _genera_estructura_posiciones(self, pPosicion):
        inserts = []
        posiciones = Posicion.objects.filter(nivel__estante__bodega=self.object)
        for nivel in Nivel.objects.filter(estante__bodega=self.object, vigente=True):
            for i in range(1, pPosicion+1):
                if not posiciones.filter(numero=i, nivel=nivel):
                    inserts.append(Posicion(numero=i, nivel=nivel))

        Posicion.objects.bulk_create(inserts)

    def _genera_estructura_niveles(self, pNiveles):
        inserts = []
        niveles = Nivel.objects.filter(estante__bodega=self.object)
        for estante in Estante.objects.filter(bodega=self.object, vigente=True):
            for i in range(1, pNiveles+1):
                if not niveles.filter(numero=i, estante=estante):
                    inserts.append(Nivel(numero=i, estante=estante))

        Nivel.objects.bulk_create(inserts)

    def _genera_estructura_estantes(self, pEstante):
        inserts = []
        estantes = Estante.objects.filter(bodega=self.object.id)

        for codigo in self._codigo_estante(pEstante):
            if not estantes.filter(codigo=codigo):
                inserts.append(Estante(codigo=codigo, bodega=self.object))

        Estante.objects.bulk_create(inserts)

    def _codigo_estante(self, numero=0):
        codigos = []
        for i in range(numero):
            c1 = i // 26
            c2 = i - c1*26
            codigos.append('{}{}'.format('' if i<26 else ascii_uppercase[c1-1],ascii_uppercase[c2]))
            i+=1
        return codigos

class Bodega_CreateView(CreateView_Login):
    permission_required = 'expedientes.add_bodega'
    model = Bodega
    form_class = Bodega_From
    template_name = 'expedientes/form.html'
    extra_context = {
        'title': _('Nueva Bodega'),
        'botones': {
            'guardar': _('Guardar'),
            'cancelar': _('Cancelar'),
        },
        'list_url': reverse_lazy('expedientes:bodega_list'),
    }

class Bodega_UpdateView(UpdateView_Login):
    permission_required = 'expedientes.change_bodega'
    model = Bodega
    form_class = Bodega_From
    template_name = 'expedientes/form.html'
    extra_context = {
        'title': _('Modificar Bodega'),
        'botones': {
            'guardar': _('Guardar'),
            'cancelar': _('Cancelar'),
        },
        'list_url': reverse_lazy('expedientes:bodega_list'),
    }

class Bodega_DeleteView(DeleteView_Login):
    permission_required = 'expedientes.delete_bodega'
    model = Bodega
    template_name = 'expedientes/confirmation_form.html'
    extra_context = {
        'title': _('Cambiar Estado Bodega'),
        'list_url': reverse_lazy('expedientes:bodega_list'),
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['botones']={
            'guardar':_('Inhabilitar') if self.object.vigente else _('Habilitar'),
            'cancelar': _('Cancelar'),
            'tag': _('danger') if self.object.vigente else _('success'),
        }
        accion = context['botones']['guardar'].lower()
        context['mensajes']={
            'confirmacion': _(f'¿Quiere {accion} el elemento indicado?'),
        }
        return context

    def get_success_url(self, *args, **kwargs):
        return super().get_success_url(self.object.vigente)


class Estante_DetailView(DetailView_Login):
    permission_required = 'expedientes.view_estante'
    model = Estante
    extra_context = {
        'title': _('Estante'),
        'sub_titulo': {
            'estructura': _('Estructura'),
        },
        'opciones':{
            'etiqueta':_('Opciones'),
            'etiquetas':_('Etiquetas'),
        },
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        
        queryset = Nivel.objects.filter(estante=self.object).order_by('numero')
        context['estructura']={
            'niveles': queryset,
            'no_columnas': queryset.count(),
            'tooltip_head': _('Nivel'),
            'tooltip_body': _('Posición'),
            'posiciones': Posicion.objects.filter(nivel__estante=self.object).order_by('numero', 'nivel__numero')
        }
        return context

class Estante_Etiqueta(DetailView_Login):
    permission_required = 'expedientes.label_estante'
    template_name = 'expedientes/etiqueta.html'
    model = Caja
    form_class = GeneraEtiquetas_Form()

    def get(self, request, *args, **kwargs):
        context = {
            'title': _('Etiqueta'),
            'object': Estante.objects.get(pk=kwargs['pk'])
        }
        arreglo = {}
        if request.GET.get('posicion'):
            cajas = list(Caja.objects.filter(posicion__nivel__estante__id=self.kwargs['pk']))
            for i in range(1, int(request.GET.get('posicion'))):
                arreglo[i] = 'ocultar'
            for caja in cajas:
                arreglo[caja] = 'mostrar'
        return render(request, self.template_name, {'arreglo': arreglo, 'form': self.form_class, 'context': context})


class Nivel_DetailView(DetailView_Login):
    permission_required = 'expedientes.view_nivel'
    model = Nivel
    extra_context = {
        'title': _('Nivel'),
        'sub_titulo': {
            'estructura': _('Estructura'),
        },
        'opciones':{
            'etiqueta':_('Opciones'),
            'etiquetas':_('Etiquetas'),
        },
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        
        queryset = Posicion.objects.filter(nivel=self.object).order_by('numero')
        context['estructura']={
            'posiciones': queryset,
            'no_columnas': queryset.count(),
            'tooltip_head': _('Posicion'),
            'tooltip_body': _('Caja'),
            'cajas': Caja.objects.filter(posicion__nivel=self.object).order_by('numero', 'posicion__numero')
        }
        return context

class Nivel_Etiqueta(DetailView_Login):
    permission_required = 'expedientes.label_nivel'
    template_name = 'expedientes/etiqueta.html'
    model = Caja
    form_class = GeneraEtiquetas_Form()

    def get(self, request, *args, **kwargs):
        context = {
            'title': _('Etiqueta'),
            'object': Nivel.objects.get(pk=kwargs['pk'])
        }
        arreglo = {}
        if request.GET.get('posicion'):
            cajas = list(Caja.objects.filter(posicion__nivel__id=self.kwargs['pk']))
            for i in range(1, int(request.GET.get('posicion'))):
                arreglo[i] = 'ocultar'
            for caja in cajas:
                arreglo[caja] = 'mostrar'
        return render(request, self.template_name, {'arreglo': arreglo, 'form': self.form_class, 'context': context})


class Posicion_DetailView(DetailView_Login):
    permission_required = 'expedientes.view_posicion'
    model = Posicion
    extra_context = {
        'title': _('Posición'),
        'sub_titulo': {
            'estructura': _('Estructura'),
        },
        'opciones':{
            'etiqueta':_('Opciones'),
            'etiquetas':_('Etiquetas'),
        },
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        
        queryset = Caja.objects.filter(posicion=self.object).order_by('numero')
        context['estructura']={
            'cajas': queryset,
            'no_columnas': queryset.count(),
            'tooltip_head': _('Caja'),
            'tooltip_body': _('Tomo'),
            'tomos': Tomo.objects.filter(caja__posicion=self.object).order_by('credito__numero', 'numero')
        }
        return context

class Posicion_Etiqueta(DetailView_Login):
    permission_required = 'expedientes.label_posicion'
    template_name = 'expedientes/etiqueta.html'
    model = Caja
    form_class = GeneraEtiquetas_Form()

    def get(self, request, *args, **kwargs):
        context = {
            'title': _('Etiqueta'),
            'object': Posicion.objects.get(pk=kwargs['pk'])
        }
        arreglo = {}
        if request.GET.get('posicion'):
            cajas = list(Caja.objects.filter(posicion__id=self.kwargs['pk']))
            for i in range(1, int(request.GET.get('posicion'))):
                arreglo[i] = 'ocultar'
            for caja in cajas:
                arreglo[caja] = 'mostrar'
        return render(request, self.template_name, {'arreglo': arreglo, 'form': self.form_class, 'context': context})


class Caja_ListView(ListView_Login):
    permission_required = 'expedientes.view_caja'
    model = Caja
    paginate_by = 15
    ordering = ['posicion__nivel__estante__codigo', 'posicion__nivel__numero', 'posicion__numero', 'numero']
    extra_context = {
        'title': _('Cajas'),
        'botones': {
            'buscar': _('Buscar'),
            'limpiar': _('Limpiar'),
        },
        'etiquetas': {
            'caja': _('Caja'),
        },
        'opciones': {
            'etiqueta': _('Opciones'),
            'inhabilitar': _('Inhabilitar'),
            'habilitar': _('Habilitar'),
        },
        'mensaje_vacio': _('No hay "Cajas" inhabilitadas'),
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.filter(vigente=False)

class Caja_DetailView(DetailView_Login):
    permission_required = 'expedientes.view_caja'
    model = Caja
    extra_context = {
        'title': _('Caja'),
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['opciones'] = {
            'etiqueta':_('Opciones'),
            'etiquetas':_('Etiquetas'),
            'accion': _('Inhabilitar') if self.object.vigente else _('Habilitar'),
            'accion_tag': 'danger' if self.object.vigente else 'success',
        }
        return context

class Caja_DeleteView(DeleteView_Login):
    permission_required = 'expedientes.delete_caja'
    model = Caja
    template_name = 'expedientes/confirmation_form.html'
    extra_context = {
        'title': _('Cambiar Estado de Caja'),
        'list_url': reverse_lazy('expedientes:index'),
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['botones']={
            'guardar':_('Inhabilitar') if self.object.vigente else _('Habilitar'),
            'cancelar': _('Cancelar'),
            'tag': _('danger') if self.object.vigente else _('success'),
        }
        accion = context['botones']['guardar'].lower()
        context['mensajes']={
            'confirmacion': _(f'¿Quiere {accion} el elemento indicado?'),
        }
        return context

    def get_success_url(self, *args, **kwargs):
        return self.object.view_url()

class Caja_Etiqueta(DetailView_Login):
    permission_required = 'expedientes.label_caja'
    template_name = 'expedientes/etiqueta.html'
    model = Caja
    form_class = GeneraEtiquetas_Form()

    def get(self, request, *args, **kwargs):
        caja = Caja.objects.get(id=self.kwargs['pk'])
        context = {
            'title': _('Etiqueta'),
            'object': caja
        }
        arreglo = {}
        if request.GET.get('posicion'):
            for i in range(1, int(request.GET.get('posicion'))):
                arreglo[i] = 'ocultar'
            arreglo[caja] = 'mostrar'
        return render(request, self.template_name, {'arreglo': arreglo, 'form': self.form_class, 'context': context})


class CargaMasiva_Form(FormView_Login):
    permission_required = 'expedientes.load_credito'
    form_class = CargaCreditos_Form
    template_name = 'expedientes/form_loadfile.html'
    success_message = _('Se cargaron los registros exitosamente')
    extra_context = {
        'title': _('Carga de Créditos'),
        'botones': {
            'guardar': _('Cargar'),
            'cancelar': _('Cancelar'),
        }
    }

    def get_success_url(self):
        return reverse_lazy('expedientes:index')

    def post(self, request, *args, **kwargs):
        if 'Cargar' in request.POST:
            form = self.get_form()
            if form.is_valid():
                return self.form_valid(form)
            else: 
                return self.form_invalid(form)

    def form_valid(self, form):
        archivo = form.cleaned_data['archivo']
        libro = openpyxl.load_workbook(archivo)
        insert_datos(datos_excel(libro[libro.sheetnames[0]]))
        return super().form_valid(form)

class Credito_DetailView(DetailView_Login):
    permission_required = 'expedientes.view_credito'
    model = Credito
    extra_context = {
        'title': _('Credito'),
        'sub_titulo': {
            'tomos': _('Tomos'),
        },
        'etiquetas':{
            'tomo':_('Cant. de Tomos'),
        },
        'opciones':{
            'etiqueta':_('Opciones'),
            'etiquetas':_('Etiquetas'),
            'agregar_tomo':_('Agregar tomo'),
            'remover_tomo':_('Remover tomo'),
            'extraer':_('Extraer'),
            'guardar':_('Guardar'),
        },
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['egreso_form']=EgresoTomo_Form()
        context['tomos'] = Tomo.objects.filter(credito=self.object, vigente=True).order_by('numero')
        return context

class Credito_Etiqueta(DetailView_Login):
    permission_required = 'expedientes.label_credito'
    template_name = 'expedientes/etiqueta.html'
    model = Tomo
    form_class = GeneraEtiquetas_Form()

    def get(self, request, *args, **kwargs):
        context = {
            'title': _('Etiqueta'),
            'object': Credito.objects.get(pk=kwargs['pk'])
        }
        arreglo = {}
        if request.GET.get('posicion'):
            tomos = list(Tomo.objects.filter(credito__id=self.kwargs['pk'], vigente=True))
            for i in range(1, int(request.GET.get('posicion'))):
                arreglo[i] = 'ocultar'
            for tomo in tomos:
                arreglo[tomo] = 'mostrar'
        return render(request, self.template_name, {'arreglo': arreglo, 'form': self.form_class, 'context': context})


class Tomo_Ingreso(FormView_Login):
    permission_required = 'expedientes.change_tomo'
    form_class = IngresoTomo_Form
    template_name = 'expedientes/form.html'
    extra_context = {
        'title': _('Ingreso a Bodega'),
        'botones': {
            'guardar': _('Guardar'),
            'cancelar': _('Cancelar'),
        },
        'list_url': reverse_lazy('expedientes:index'),
    }
    success_url = reverse_lazy('expedientes:index')
    success_message =''

    def post(self, request, *args, **kwargs):
        if 'Guardar' in request.POST:
            form = self.get_form()
            if form.is_valid():
                return self.form_valid(form)
            else: 
                return self.form_invalid(form)
    
    def form_valid(self, form, *args, **kwargs):
        tomo = form.cleaned_data['tomo']
        caja = form.cleaned_data['caja']
        comentario = form.cleaned_data['comentario']

        try:
            tomo_qs = Tomo.objects.get(credito__numero=tomo[0], numero=tomo[1])
            caja_qs = Caja.objects.get(posicion__nivel__estante__bodega__codigo=caja[0].upper(), 
                posicion__nivel__estante__codigo=caja[1].upper(), posicion__nivel__numero=caja[2], 
                posicion__numero=caja[3], numero=caja[4])

            if not caja_qs.vigente:
                messages.warning(self.request, _('La caja no se encuentra habilitada'))
            elif not tomo_qs.caja:
                tomo_qs.comentario = comentario
                tomo_qs.vigente = True
                tomo_qs.caja = caja_qs
                tomo_qs.usuario = self.request.user
                tomo_qs.save()
                messages.success(self.request, _('Se guardo el tomo ')+f'{tomo[0]}-{tomo[1]}')

                return redirect(tomo_qs.credito.view_url())
            else:
                messages.warning(self.request, _('El tomo se encuentra asignado a: ')+f'{tomo_qs.caja}')
        except:
            messages.warning(self.request, _('Tomo o caja no encontrado ')+f'{tomo[0]}-{tomo[1]}')
        
        return super().form_valid(form)

class Tomo_Etiqueta(DetailView_Login):
    permission_required = 'expedientes.label_tomo'
    template_name = 'expedientes/etiqueta.html'
    model = Tomo
    form_class = GeneraEtiquetas_Form()

    def get(self, request, *args, **kwargs):
        context = {
            'title': _('Etiqueta'),
            'object': Tomo.objects.get(pk=kwargs['pk'])
        }
        arreglo = {}
        if request.GET.get('posicion'):
            tomos = list(Tomo.objects.filter(id=self.kwargs['pk']))
            for i in range(1, int(request.GET.get('posicion'))):
                arreglo[i] = 'ocultar'
            for tomo in tomos:
                arreglo[tomo] = 'mostrar'
        return render(request, self.template_name, {'arreglo': arreglo, 'form': self.form_class, 'context': context})

class ProcesaEgresos(TemplateView_Login):
    permission_required = 'expedientes.change_tomo'
    template_name = 'expedientes/tomo_list.html'
    extra_context = {
        'title': 'Salida de Tomos',
        'opciones': {
            'etiqueta': _('Opciones'),
            'eliminar': _('Eliminar'),
            'trasladar': _('Trasladar'),
            'egresar': _('Egresar'),
        },
        'sub_titulo': {
            'traslado':_('Traslado de bodega'),
            'egreso': _('Egreso de bodega'),
        },
        'mensaje_vacio':_('No hay tomos para enviar')
    }

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        tomos = self.request.session.get('extraer_tomos', [])
        tomo_qs = Tomo.objects.filter(id__in=tomos).order_by('credito__numero', 'numero')
        context['object_list'] = tomo_qs
        context['traslado_form'] = TrasladoTomos_Form()
        context['egreso_form'] = SalidaTomos_Form()
        return context


def Credito_Search(request):
    numero = request.GET['numero'].replace(' ','').split('-')
    credito = Credito.objects.filter(numero=numero[0])
    if credito:
        return redirect(credito[0].view_url())
    elif len(numero)>0:
        messages.warning(request, _('No se encontró el crédito: ')+numero[0])

    #return render(request, 'expedientes/index.html')
    return redirect(reverse_lazy('expedientes:index'))
    
def Tomo_Opera(request):
    credito = Credito.objects.get(id=request.POST['credito'])
    tomos = Tomo.objects.filter(credito=credito)
    if 'agregar.x' in request.POST:
        if not tomos:
            Tomo(numero=1, credito=credito, comentario='Tomo habilitado', 
                usuario=request.user).save()
        elif tomos.filter(vigente=False):
            #tomo minimo inhabilitado
            tomo = tomos.filter(vigente=False).order_by('numero')[0]
            tomo.usuario=request.user
            tomo.vigente=True
            tomo.comentario='Tomo habilitado'
            tomo.save()
        else:
            num = tomos.aggregate(Max('numero'))['numero__max']+1
            Tomo(numero=num, credito=credito, comentario='Tomo habilitado', 
                usuario=request.user).save()
    else: #elif 'remover.x' in request.POST:
        if not tomos or not tomos.filter(vigente=True):
            messages.warning(request, _('No hay tomos para deshabilitar en el crédito ')+ credito.numero)
        else:
            #tomo máximo habilitado
            tomo = tomos.filter(vigente=True).order_by('-numero')[0]
            tomo.usuario=request.user
            tomo.caja = None
            tomo.vigente=False
            tomo.comentario='Tomo inhabilitado'
            tomo.save()
    return redirect(credito.view_url())

def Tomo_CargarEgreso(request):
    '''
        Agrega tomos a a lista que se extraera de la bodega.
    '''
    #del request.session['extraer_tomos']
    if 'Guardar' in request.POST:
        extraer_tomos=request.session['extraer_tomos'] if 'extraer_tomos' in request.session else []            
        tomoid = request.POST['tomo-id']
        tomo = request.POST['tomo'].replace(' ', '').split('-')
        try: 
            if len(tomo)==2:
                t = Tomo.objects.get(credito__numero=tomo[0], numero=int(tomo[1]))
                if t and str(t.id)==tomoid:
                    if not tomoid in extraer_tomos:
                        extraer_tomos.append(tomoid)
                        request.session['extraer_tomos']=extraer_tomos
                        messages.success(request, _('Tomo agregado a la lista'))
                    else:
                        messages.warning(request, _('Tomo agregado previamente'))
                else:
                    messages.warning(request, _('Tomo no corresponde al seleccionado'))
            else:
                messages.warning(request, _('Tomo mal ingresado (longitud)'))
        except:
            messages.warning(request, _('Tomo mal ingresado o no existe'))
        finally:
            return redirect(Tomo.objects.get(id=tomoid).credito.view_url())

def Tomo_EgresoRemover(request, pk):
    '''
        Elimina tomos del listado a trasladar/egresar
    '''
    try: 
        extraer_tomos = request.session['extraer_tomos']
        extraer_tomos.remove(str(pk))
        request.session['extraer_tomos'] = extraer_tomos
    finally:
        return redirect(Tomo.envio_url())
    
def Tomo_Trasladar(request):
    if request.method=='POST':
        try:
            form = TrasladoTomos_Form(request.POST)
            if form.is_valid():
                bodega = form.cleaned_data['bodega_envio']
                comentario = f'Traslado a {bodega}\n{form.cleaned_data["comentario"]}'
                tomos = Tomo.objects.filter(id__in=request.session['extraer_tomos'])
                context = {
                    'title': _('Traslado'),
                    'message': _(f'Se han enviado, a {bodega}, los siguientes tomos:'),
                    'object_list': {tomo for tomo in tomos},
                }
                tomos.update(comentario=comentario, caja=None)
                del request.session['extraer_tomos']
                messages.success(request, _('Tomos egresados por traslado'))
                correo = crea_correo('Traslado de Expedientes', request.user.email, [bodega.encargado.email], 'mails/egresos.html', context)
                envia_correo(correo)
        finally:
            return redirect(Tomo.envio_url())

def Tomo_Egresar(request):
    if request.method=='POST':
        try:
            form = SalidaTomos_Form(request.POST)
            if form.is_valid():
                correo_form = form.cleaned_data["correo"]

                comentario_final = f'Fecha: \t\t{datetime.today().strftime("%d-%m-%Y")}\n'
                comentario_final += f'Codigo: \t{form.cleaned_data["codigo"]}\n'
                comentario_final += f'Nombre: \t{form.cleaned_data["nombre"]}\n'
                comentario_final += f'Extension: \t{form.cleaned_data["extension"]}\n'
                comentario_final += f'Correo: \t{correo_form}\n'
                comentario_final += f'Gerencia: \t{form.cleaned_data["gerencia"]}\n'
                comentario_final += f'Comentario: \t{form.cleaned_data["comentario"]}'

                tomos = Tomo.objects.filter(id__in=request.session['extraer_tomos'])
                context = {
                    'title': _('Egreso por Solicitud'),
                    'message': _(f'Se han entregado, los siguientes tomos:'),
                    'object_list': {tomo for tomo in tomos},
                    'comentario': comentario_final,
                }
                tomos.update(comentario=comentario_final, caja=None)
                del request.session['extraer_tomos']
                messages.success(request, _('Tomos egresados por solicitud'))
                correo = crea_correo('Egreso por solicitud', request.user.email, [request.user.email,correo_form], 'mails/egresos.html', context)
                envia_correo(correo)
        finally:
            return redirect(Tomo.envio_url())



##########################################################################
#
##########################################################################
def crea_correo(subject, from_email, to_email, template, context):
    template = get_template(template)
    content = template.render(context)

    mail = EmailMultiAlternatives(
        subject=subject,
        body='',
        from_email=from_email,
        to=to_email,
        cc=[],
        )

    mail.attach_alternative(content, 'text/html')
    return mail

def envia_correo(mail):
    thread = threading.Thread(
        mail.send(fail_silently=False)
    )
    thread.start()
    
##########################################################################
#
##########################################################################
def datos_excel(hoja):    
    orden = {}
    for i in range(len(hoja[1])):
        orden[hoja[1][i].value]=i

    clientes = []
    oficinas = []
    monedas = []
    productos = []
    creditos = []
    for fila in hoja.iter_rows(min_row=2):
        mis     = fila[orden['Cod_Cliente']].value
        cliente = fila[orden['Cliente']].value
        if not any(mis in sublist for sublist in clientes):
            clientes.append([mis, cliente])

        cod_ofi = fila[orden['Cod_ofi']].value
        oficina = fila[orden['Oficina']].value 
        if not any(cod_ofi in sublist for sublist in oficinas):
            oficinas.append([cod_ofi, oficina])

        moneda  = fila[orden['Moneda']].value
        if not moneda in monedas:
            monedas.append(moneda)

        producto = fila[orden['Producto']].value
        if not producto in productos:
            productos.append(producto)

        credito = fila[orden['Credito']].value
        fecha   = datetime.strptime(fila[orden['Fecha_Ini']].value, "%d/%m/%Y").strftime("%Y-%m-%d")
        monto   = fila[orden['Monto']].value
        if not any(credito in sublist for sublist in creditos):
            creditos.append([credito, fecha, monto, mis, cod_ofi, moneda, producto])
        
    return {'clientes': clientes, 'oficinas': oficinas, 'monedas': monedas, 
        'productos': productos,'creditos': creditos}

def insert_datos(datos):
    clientes = insert_clientes(datos['clientes'])
    oficinas = insert_oficinas(datos['oficinas'])
    monedas = insert_monedas(datos['monedas'])
    productos = insert_productos(datos['productos'])
    creditos = insert_creditos(datos['creditos'])

def insert_clientes(datos):
    clientes = []
    queryset = Cliente.objects.filter(codigo__in=[int(dato[0]) for dato in datos])
    for dato in datos:
        if not queryset.filter(codigo=int(dato[0])):
            clientes.append(Cliente(codigo=int(dato[0]), nombre=dato[1]))
    Cliente.objects.bulk_create(clientes)

def insert_oficinas(datos):
    oficinas = []
    queryset = Oficina.objects.filter(numero__in=[int(dato[0]) for dato in datos])
    for dato in datos:
        if not queryset.filter(numero=int(dato[0])):
            oficinas.append(Oficina(numero=int(dato[0]), descripcion=dato[1]))
    Oficina.objects.bulk_create(oficinas)

def insert_monedas(datos):
    monedas = []
    queryset = Moneda.objects.filter(descripcion__in=datos)
    for dato in datos:
        if not queryset.filter(descripcion=dato):
            monedas.append(Moneda(descripcion=dato))
    Moneda.objects.bulk_create(monedas)

def insert_productos(datos):
    productos = []
    queryset = Producto.objects.filter(descripcion__in=datos)
    for dato in datos:
        if not queryset.filter(descripcion=dato):
            productos.append(Producto(descripcion=dato))
    Producto.objects.bulk_create(productos)

def insert_creditos(datos):
    #creditos.append([credito, fecha, monto, mis, cod_ofi, moneda, producto])
    queryset = Credito.objects.filter(numero__in=[dato[0] for dato in datos])
    clientes = Cliente.objects.filter(codigo__in=[int(dato[3]) for dato in datos])
    oficinas = Oficina.objects.filter(numero__in=[int(dato[4]) for dato in datos])
    monedas = Moneda.objects.filter(descripcion__in=[dato[5] for dato in datos])
    productos = Producto.objects.filter(descripcion__in=[dato[6] for dato in datos])

    creditos = []
    for dato in datos:
        if not queryset.filter(numero=dato[0]):
            creditos.append(Credito(numero=dato[0], fecha_concesion=dato[1], 
                monto=dato[2], cliente=clientes.get(codigo=int(dato[3])), 
                moneda=monedas.get(descripcion=dato[5]), 
                oficina=oficinas.get(numero=int(dato[4])), 
                producto=productos.get(descripcion=dato[6])))
    Credito.objects.bulk_create(creditos)